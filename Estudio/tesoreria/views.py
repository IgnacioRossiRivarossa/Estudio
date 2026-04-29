import json
import logging
from decimal import Decimal, InvalidOperation
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.db.models import Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_POST

from cotizaciones.services import get_dolares
from .forms import (
    BancoForm, CajaForm, FCIForm, MonedaExtranjeraForm,
    PlazoFijoForm, TituloONForm, ValorADepositarForm,
)
from .models import (
    Banco, Caja, FCI, MonedaExtranjera, PlazoFijo, TituloON,
    ValorADepositar, ValorADepositarEmpresa,
)

logger = logging.getLogger(__name__)

def _check_rol(user):
    if user.rol not in ('Administrador', 'Colaborador'):
        raise PermissionDenied

def _get_dolar_mep():
    """Obtiene cotización Dólar MEP reutilizando el módulo cotizaciones."""
    dolares = get_dolares()
    for d in dolares:
        if d.get('casa') == 'bolsa' or 'mep' in d.get('nombre', '').lower():
            return d
    return None

def _safe_sum(queryset, field):
    return queryset.aggregate(total=Sum(field))['total'] or Decimal('0')


@login_required
def dashboard(request):
    _check_rol(request.user)

    # Dólar MEP
    dolar_mep = _get_dolar_mep()
    dolar_mep_venta = Decimal(str(dolar_mep.get('venta', 0))) if dolar_mep else Decimal('0')

    # Caja
    cajas = Caja.objects.all()
    total_caja = _safe_sum(cajas, 'saldo')
    total_caja_sem_ant = _safe_sum(cajas, 'saldo_sem_ant')

    # Bancos
    bancos = Banco.objects.all()
    total_bancos_pesos = _safe_sum(bancos, 'saldo_cuenta_corriente')
    total_bancos_usd = _safe_sum(bancos, 'saldo_usd')
    total_bancos_sem_ant_pesos = _safe_sum(bancos, 'saldo_sem_ant_pesos')
    total_bancos_sem_ant_usd = _safe_sum(bancos, 'saldo_usd_sem_ant')

    # Moneda Extranjera
    monedas = MonedaExtranjera.objects.all()
    total_me = _safe_sum(monedas, 'saldo_dolares')
    total_me_sem_ant = _safe_sum(monedas, 'saldo_dolares_sem_ant')

    # Valores a Depositar
    total_vad = _safe_sum(ValorADepositar.objects.all(), 'monto')
    total_vad_sem_ant = _safe_sum(ValorADepositarEmpresa.objects.all(), 'saldo_sem_ant')

    # Inversiones
    total_pf = _safe_sum(PlazoFijo.objects.all(), 'monto_invertido')
    fcis = FCI.objects.all()
    total_fci = _safe_sum(fcis, 'saldo')
    total_fci_sem_ant = _safe_sum(fcis, 'saldo_sem_ant')
    titulos = TituloON.objects.all()
    total_titulos_pesos = _safe_sum(titulos, 'saldo_pesos_actual')
    total_titulos_usd = _safe_sum(titulos, 'saldo_usd_actual')
    total_titulos_pesos_sem_ant = _safe_sum(titulos, 'saldo_pesos_sem_ant')
    total_titulos_usd_sem_ant = _safe_sum(titulos, 'saldo_usd_sem_ant')
    total_inversiones = total_pf + total_fci + total_titulos_pesos
    total_inversiones_sem_ant = total_fci_sem_ant + total_titulos_pesos_sem_ant

    # Caja y Bancos total
    total_caja_bancos_pesos = total_caja + total_bancos_pesos
    total_caja_bancos_usd = total_bancos_usd + total_me
    total_caja_bancos_sem_ant_pesos = total_caja_sem_ant + total_bancos_sem_ant_pesos
    total_caja_bancos_sem_ant_usd = total_bancos_sem_ant_usd + total_me_sem_ant

    context = {
        'dolar_mep': dolar_mep,
        'dolar_mep_venta': dolar_mep_venta,
        'total_caja': total_caja,
        'total_caja_sem_ant': total_caja_sem_ant,
        'total_bancos_pesos': total_bancos_pesos,
        'total_bancos_usd': total_bancos_usd,
        'total_bancos_sem_ant_pesos': total_bancos_sem_ant_pesos,
        'total_bancos_sem_ant_usd': total_bancos_sem_ant_usd,
        'total_me': total_me,
        'total_me_sem_ant': total_me_sem_ant,
        'total_vad': total_vad,
        'total_vad_sem_ant': total_vad_sem_ant,
        'total_pf': total_pf,
        'total_fci': total_fci,
        'total_fci_sem_ant': total_fci_sem_ant,
        'total_titulos_pesos': total_titulos_pesos,
        'total_titulos_usd': total_titulos_usd,
        'total_titulos_pesos_sem_ant': total_titulos_pesos_sem_ant,
        'total_titulos_usd_sem_ant': total_titulos_usd_sem_ant,
        'total_inversiones': total_inversiones,
        'total_inversiones_sem_ant': total_inversiones_sem_ant,
        'total_caja_bancos_pesos': total_caja_bancos_pesos,
        'total_caja_bancos_usd': total_caja_bancos_usd,
        'total_caja_bancos_sem_ant_pesos': total_caja_bancos_sem_ant_pesos,
        'total_caja_bancos_sem_ant_usd': total_caja_bancos_sem_ant_usd,
        'vista_activa': 'dashboard',
    }
    return render(request, 'tesoreria/dashboard.html', context)

@login_required
def caja_bancos(request):
    _check_rol(request.user)

    cajas = Caja.objects.all()
    bancos = Banco.objects.all()
    monedas = MonedaExtranjera.objects.all()

    total_caja = _safe_sum(cajas, 'saldo')
    total_caja_sem_ant = _safe_sum(cajas, 'saldo_sem_ant')
    total_bancos_cc = _safe_sum(bancos, 'saldo_cuenta_corriente')
    total_bancos_chq_acred = _safe_sum(bancos, 'cheque_pendiente_acreditar')
    total_bancos_chq_deb = _safe_sum(bancos, 'cheque_pendiente_debito')
    total_bancos_usd = _safe_sum(bancos, 'saldo_usd')
    total_bancos_sem_ant_pesos = _safe_sum(bancos, 'saldo_sem_ant_pesos')
    total_bancos_sem_ant_usd = _safe_sum(bancos, 'saldo_usd_sem_ant')
    total_me = _safe_sum(monedas, 'saldo_dolares')
    total_me_sem_ant = _safe_sum(monedas, 'saldo_dolares_sem_ant')

    context = {
        'cajas': cajas,
        'bancos': bancos,
        'monedas': monedas,
        'total_caja': total_caja,
        'total_caja_sem_ant': total_caja_sem_ant,
        'total_bancos_cc': total_bancos_cc,
        'total_bancos_chq_acred': total_bancos_chq_acred,
        'total_bancos_chq_deb': total_bancos_chq_deb,
        'total_bancos_usd': total_bancos_usd,
        'total_bancos_sem_ant_pesos': total_bancos_sem_ant_pesos,
        'total_bancos_sem_ant_usd': total_bancos_sem_ant_usd,
        'total_me': total_me,
        'total_me_sem_ant': total_me_sem_ant,
        'vista_activa': 'caja_bancos',
    }
    return render(request, 'tesoreria/caja_bancos.html', context)

@login_required
def valores_a_depositar(request):
    _check_rol(request.user)

    valores = ValorADepositar.objects.all()

    # Agrupar por empresa
    total_por_empresa = (
        valores.values('empresa')
        .annotate(total=Sum('monto'))
        .order_by('empresa')
    )

    # Agrupar por mes de vencimiento
    total_por_mes = (
        valores.values('mes_vencimiento', 'anio_vencimiento')
        .annotate(total=Sum('monto'))
        .order_by('anio_vencimiento', 'mes_vencimiento')
    )

    total_vad = _safe_sum(valores, 'monto')

    # Datos para Chart.js
    chart_labels = [f"{r['mes_vencimiento']} {r['anio_vencimiento']}" for r in total_por_mes]
    chart_values = [float(r['total']) for r in total_por_mes]

    vad_sem_ant_map = {
        obj.empresa: obj.saldo_sem_ant
        for obj in ValorADepositarEmpresa.objects.all()
    }

    context = {
        'valores': valores,
        'total_por_empresa': total_por_empresa,
        'total_por_mes': total_por_mes,
        'total_vad': total_vad,
        'chart_labels': json.dumps(chart_labels),
        'chart_values': json.dumps(chart_values),
        'vad_sem_ant_map': vad_sem_ant_map,
        'vista_activa': 'valores_a_depositar',
    }
    return render(request, 'tesoreria/valores_a_depositar.html', context)

@login_required
def inversiones(request):
    _check_rol(request.user)

    plazos_fijos = PlazoFijo.objects.all()
    fcis = FCI.objects.all()
    titulos = TituloON.objects.all()

    total_pf = _safe_sum(plazos_fijos, 'monto_invertido')
    total_fci = _safe_sum(fcis, 'saldo')
    total_titulos_pesos = _safe_sum(titulos, 'saldo_pesos_actual')
    total_titulos_usd = _safe_sum(titulos, 'saldo_usd_actual')
    total_inversiones = total_pf + total_fci + total_titulos_pesos

    context = {
        'plazos_fijos': plazos_fijos,
        'fcis': fcis,
        'titulos': titulos,
        'total_pf': total_pf,
        'total_fci': total_fci,
        'total_titulos_pesos': total_titulos_pesos,
        'total_titulos_usd': total_titulos_usd,
        'total_inversiones': total_inversiones,
        'vista_activa': 'inversiones',
    }
    return render(request, 'tesoreria/inversiones.html', context)

@login_required
@require_POST
def actualizar_precios_titulos(request):
    _check_rol(request.user)

    titulos = TituloON.objects.all()
    if not titulos.exists():
        return JsonResponse({"actualizados": 0, "errores": ["No hay títulos cargados."]})

    # Separar los que tienen precio manual (no se consultan en IOL)
    titulos_iol = [t for t in titulos if not t.precio_manual]

    tickers_a_consultar = set()
    for t in titulos_iol:
        es_panama = "panama" in t.nombre.lower()
        ticker = t.ticker.upper()
        if es_panama:
            # ON Panama: ticker almacenado es el completo terminado en C (dólar cable)
            tickers_a_consultar.add(ticker)
        elif t.tipo == 'ON':
            # ON normal: ticker es la base (ej: BYCH), o ya trae O/D → normalizar
            base = ticker[:-1] if ticker.endswith(('O', 'D')) else ticker
            tickers_a_consultar.add(base + "O")  # pesos
            tickers_a_consultar.add(base + "D")  # dólares
        else:
            if ticker.endswith('D'):
                # Caso D: ticker ya está en dólares (ej: TX26D)
                tickers_a_consultar.add(ticker)        # dólares
                tickers_a_consultar.add(ticker[:-1])   # pesos (ej: TX26)
            else:
                # Caso C: ticker en pesos (ej: TX26), dólares = ticker + D
                tickers_a_consultar.add(ticker)        # pesos
                tickers_a_consultar.add(ticker + "D")  # dólares

    from .iol_client import IOLClient, consultar_titulos_paralelo

    iol = IOLClient()
    resultado = consultar_titulos_paralelo(iol, list(tickers_a_consultar))
    cierres = resultado["cierres"]
    errores = resultado["errores"]

    dolar_mep = _get_dolar_mep()
    dolar_mep_venta = Decimal(str(dolar_mep.get('venta', 0))) if dolar_mep else Decimal('0')

    manuales_omitidos = len(titulos) - len(titulos_iol)
    actualizados = 0
    for t in titulos_iol:
        es_panama = "panama" in t.nombre.lower()
        ticker = t.ticker.upper()
        try:
            if es_panama:
                # Panama ON: ticker termina en C → valor en USD
                # saldo_usd = cuotapartes * (cierre_usd / 100)
                # saldo_pesos = saldo_usd * dolar_mep
                cierre_usd = cierres.get(ticker)
                t.valor_cierre_pesos = None
                if cierre_usd is not None:
                    t.valor_cierre_usd = Decimal(str(cierre_usd))
                    t.saldo_usd_actual = t.cuotapartes_actual * (t.valor_cierre_usd / Decimal('100'))
                    t.saldo_pesos_actual = t.saldo_usd_actual * dolar_mep_venta
                else:
                    t.valor_cierre_usd = None
                    t.saldo_usd_actual = None
                    t.saldo_pesos_actual = None
            elif t.tipo == 'ON':
                # ON normal: base + O (pesos), base + D (dólares)
                # saldo = cuotapartes * (cierre / 100)
                base = ticker[:-1] if ticker.endswith(('O', 'D')) else ticker
                cierre_pesos = cierres.get(base + "O")
                cierre_usd   = cierres.get(base + "D")
                if cierre_pesos is not None:
                    t.valor_cierre_pesos = Decimal(str(cierre_pesos))
                    t.saldo_pesos_actual = t.cuotapartes_actual * (t.valor_cierre_pesos / Decimal('100'))
                else:
                    t.valor_cierre_pesos = None
                    t.saldo_pesos_actual = None
                if cierre_usd is not None:
                    t.valor_cierre_usd = Decimal(str(cierre_usd))
                    t.saldo_usd_actual = t.cuotapartes_actual * (t.valor_cierre_usd / Decimal('100'))
                else:
                    t.valor_cierre_usd = None
                    t.saldo_usd_actual = None
            elif ticker.endswith('D'):
                # Caso D: ticker ya está en dólares (ej: TX26D)
                # saldo_usd = cuotapartes * (cierre_usd / 100)
                # saldo_pesos = cuotapartes * (cierre_pesos / 100), o saldo_usd * MEP si no existe
                cierre_usd   = cierres.get(ticker)
                cierre_pesos = cierres.get(ticker[:-1])
                if cierre_usd is not None:
                    t.valor_cierre_usd = Decimal(str(cierre_usd))
                    t.saldo_usd_actual = t.cuotapartes_actual * (t.valor_cierre_usd / Decimal('100'))
                else:
                    t.valor_cierre_usd = None
                    t.saldo_usd_actual = None
                if cierre_pesos is not None:
                    t.valor_cierre_pesos = Decimal(str(cierre_pesos))
                    t.saldo_pesos_actual = t.cuotapartes_actual * (t.valor_cierre_pesos / Decimal('100'))
                else:
                    t.valor_cierre_pesos = None
                    # Sin ticker pesos: equivalente = saldo_usd * dolar_mep
                    if t.saldo_usd_actual is not None and dolar_mep_venta > 0:
                        t.saldo_pesos_actual = t.saldo_usd_actual * dolar_mep_venta
                    else:
                        t.saldo_pesos_actual = None
            else:
                # Caso C: ticker en pesos (ej: TX26), dólares = ticker + D
                # saldo = cuotapartes * (cierre / 100)
                cierre_pesos = cierres.get(ticker)
                cierre_usd   = cierres.get(ticker + "D")
                if cierre_pesos is not None:
                    t.valor_cierre_pesos = Decimal(str(cierre_pesos))
                    t.saldo_pesos_actual = t.cuotapartes_actual * (t.valor_cierre_pesos / Decimal('100'))
                else:
                    t.valor_cierre_pesos = None
                    t.saldo_pesos_actual = None
                if cierre_usd is not None:
                    t.valor_cierre_usd = Decimal(str(cierre_usd))
                    t.saldo_usd_actual = t.cuotapartes_actual * (t.valor_cierre_usd / Decimal('100'))
                else:
                    t.valor_cierre_usd = None
                    # Sin ticker USD: equivalente = saldo_pesos / dolar_mep
                    if t.saldo_pesos_actual is not None and dolar_mep_venta > 0:
                        t.saldo_usd_actual = t.saldo_pesos_actual / dolar_mep_venta
                    else:
                        t.saldo_usd_actual = None

            t.save(update_fields=[
                "valor_cierre_pesos", "valor_cierre_usd",
                "saldo_pesos_actual", "saldo_usd_actual",
            ])
            actualizados += 1

        except (InvalidOperation, Exception) as e:
            errores.append(f"{t.ticker}: error al guardar — {e}")

    return JsonResponse({
        "actualizados": actualizados,
        "errores": errores,
        "manuales_omitidos": manuales_omitidos,
    })

@login_required
@require_POST
def actualizar_semana_titulos(request):
    _check_rol(request.user)

    titulos = TituloON.objects.all()
    actualizados = 0
    errores = []

    for t in titulos:
        try:
            t.cuotapartes_sem_ant = t.cuotapartes_actual
            t.saldo_pesos_sem_ant = t.saldo_pesos_actual
            t.saldo_usd_sem_ant = t.saldo_usd_actual
            t.save(update_fields=[
                'cuotapartes_sem_ant', 'saldo_pesos_sem_ant', 'saldo_usd_sem_ant',
            ])
            actualizados += 1
        except Exception as e:
            errores.append(f"{t.ticker}: {e}")

    return JsonResponse({"actualizados": actualizados, "errores": errores})

@login_required
@require_POST
def actualizar_semana_fci(request):
    _check_rol(request.user)

    fcis = FCI.objects.all()
    actualizados = 0
    errores = []

    for f in fcis:
        try:
            f.cuotapartes_sem_ant = f.cuotapartes
            f.saldo_sem_ant = f.saldo
            f.save(update_fields=['cuotapartes_sem_ant', 'saldo_sem_ant'])
            actualizados += 1
        except Exception as e:
            errores.append(f"{f.nombre}: {e}")

    return JsonResponse({"actualizados": actualizados, "errores": errores})

@login_required
@require_POST
def actualizar_semana_caja(request):
    _check_rol(request.user)
    cajas = Caja.objects.all()
    for c in cajas:
        c.saldo_sem_ant = c.saldo
        c.save(update_fields=['saldo_sem_ant'])
    return JsonResponse({"actualizados": cajas.count()})


@login_required
@require_POST
def actualizar_semana_me(request):
    _check_rol(request.user)
    monedas = MonedaExtranjera.objects.all()
    for m in monedas:
        m.saldo_dolares_sem_ant = m.saldo_dolares
        m.save(update_fields=['saldo_dolares_sem_ant'])
    return JsonResponse({"actualizados": monedas.count()})


@login_required
@require_POST
def actualizar_semana_bancos(request):
    _check_rol(request.user)
    bancos = Banco.objects.all()
    for b in bancos:
        b.saldo_sem_ant_pesos = (
            (b.saldo_cuenta_corriente or Decimal('0'))
            + (b.cheque_pendiente_acreditar or Decimal('0'))
            + (b.cheque_pendiente_debito or Decimal('0'))
        )
        b.saldo_usd_sem_ant = b.saldo_usd
        b.save(update_fields=['saldo_sem_ant_pesos', 'saldo_usd_sem_ant'])
    return JsonResponse({"actualizados": bancos.count()})


@login_required
@require_POST
def actualizar_semana_vad(request):
    _check_rol(request.user)
    totales = ValorADepositar.objects.values('empresa').annotate(total=Sum('monto'))
    for row in totales:
        obj, _ = ValorADepositarEmpresa.objects.get_or_create(empresa=row['empresa'])
        obj.saldo_sem_ant = row['total']
        obj.save(update_fields=['saldo_sem_ant'])
    return JsonResponse({"actualizados": totales.count()})


@login_required
def exportar_excel(request):
    _check_rol(request.user)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── Paleta ──────────────────────────────────────────────────────────────
    COLOR_HEADER   = '8C4F9F'   # púrpura — cabecera de columnas
    COLOR_SECCION  = '3D2B4F'   # púrpura oscuro — título de sección
    COLOR_TOTAL    = 'F3EEF7'   # lavanda suave — fila total
    COLOR_SEM_ANT  = 'E8E8E8'   # gris claro — columnas semana anterior

    font_header  = Font(bold=True, color='FFFFFF', size=10)
    font_seccion = Font(bold=True, color='FFFFFF', size=11)
    font_total   = Font(bold=True, size=10)
    font_normal  = Font(size=10)

    fill_header  = PatternFill('solid', fgColor=COLOR_HEADER)
    fill_seccion = PatternFill('solid', fgColor=COLOR_SECCION)
    fill_total   = PatternFill('solid', fgColor=COLOR_TOTAL)
    fill_sem_ant = PatternFill('solid', fgColor=COLOR_SEM_ANT)

    aln_center = Alignment(horizontal='center', vertical='center')
    aln_right  = Alignment(horizontal='right',  vertical='center')
    aln_left   = Alignment(horizontal='left',   vertical='center')

    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def fmt(val):
        """Decimal/None → float o None para Excel."""
        if val is None:
            return None
        try:
            return float(val)
        except (TypeError, ValueError):
            return None

    def escribir_seccion(ws, titulo, headers, filas, sem_ant_cols=None):
        """
        Escribe un bloque: título + cabecera + filas.
        sem_ant_cols: set de índices (1-based) con columnas de semana anterior.
        Devuelve la fila siguiente disponible.
        """
        sem_ant_cols = sem_ant_cols or set()

        # Título de sección
        ws.append([titulo])
        fila_titulo = ws.max_row
        cell = ws.cell(fila_titulo, 1)
        cell.font = font_seccion
        cell.fill = fill_seccion
        cell.alignment = aln_left
        ws.merge_cells(
            start_row=fila_titulo, start_column=1,
            end_row=fila_titulo, end_column=len(headers)
        )

        # Cabecera
        ws.append(headers)
        fila_cab = ws.max_row
        for col_idx, _ in enumerate(headers, 1):
            c = ws.cell(fila_cab, col_idx)
            c.font = font_header
            c.fill = fill_header
            c.alignment = aln_center
            c.border = border

        # Filas de datos
        for fila in filas:
            ws.append(fila)
            fila_num = ws.max_row
            for col_idx, val in enumerate(fila, 1):
                c = ws.cell(fila_num, col_idx)
                c.font = font_normal
                c.border = border
                if col_idx in sem_ant_cols:
                    c.fill = fill_sem_ant
                if isinstance(val, float):
                    c.alignment = aln_right
                    c.number_format = '#,##0.00'
                else:
                    c.alignment = aln_left

    def escribir_total(ws, n_cols, valores_por_col, sem_ant_cols=None):
        """Agrega fila TOTAL con los valores en las columnas indicadas."""
        sem_ant_cols = sem_ant_cols or set()
        fila = ['TOTAL'] + [None] * (n_cols - 1)
        for col_idx, val in valores_por_col.items():
            fila[col_idx - 1] = val
        ws.append(fila)
        fila_num = ws.max_row
        for col_idx in range(1, n_cols + 1):
            c = ws.cell(fila_num, col_idx)
            c.font = font_total
            c.fill = fill_total
            c.border = border
            if col_idx in sem_ant_cols:
                c.fill = PatternFill('solid', fgColor='D5C8E0')
            v = ws.cell(fila_num, col_idx).value
            if isinstance(v, float):
                c.alignment = aln_right
                c.number_format = '#,##0.00'
            else:
                c.alignment = aln_left

    def ajustar_columnas(ws, min_width=12):
        for col in ws.columns:
            max_len = min_width
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)) + 4)
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len, 40)

    # ── Datos ────────────────────────────────────────────────────────────────
    cajas   = Caja.objects.all().order_by('empresa')
    bancos  = Banco.objects.all().order_by('nombre')
    monedas = MonedaExtranjera.objects.all().order_by('empresa')

    valores = ValorADepositar.objects.all()
    total_por_empresa = (
        valores.values('empresa').annotate(total=Sum('monto')).order_by('empresa')
    )
    total_por_mes = (
        valores.values('mes_vencimiento', 'anio_vencimiento')
        .annotate(total=Sum('monto'))
        .order_by('anio_vencimiento', 'mes_vencimiento')
    )
    vad_sem_ant_map = {
        obj.empresa: obj.saldo_sem_ant
        for obj in ValorADepositarEmpresa.objects.all()
    }

    plazos_fijos = PlazoFijo.objects.all().order_by('banco')
    fcis         = FCI.objects.all().order_by('nombre')
    titulos      = TituloON.objects.all().order_by('nombre')

    # Totales
    t_caja             = _safe_sum(cajas,   'saldo')
    t_caja_sem         = _safe_sum(cajas,   'saldo_sem_ant')
    t_bancos_cc        = _safe_sum(bancos,  'saldo_cuenta_corriente')
    t_bancos_chq_acred = _safe_sum(bancos,  'cheque_pendiente_acreditar')
    t_bancos_chq_deb   = _safe_sum(bancos,  'cheque_pendiente_debito')
    t_bancos_usd       = _safe_sum(bancos,  'saldo_usd')
    t_bancos_sem_pesos = _safe_sum(bancos,  'saldo_sem_ant_pesos')
    t_bancos_sem_usd   = _safe_sum(bancos,  'saldo_usd_sem_ant')
    t_me               = _safe_sum(monedas, 'saldo_dolares')
    t_me_sem           = _safe_sum(monedas, 'saldo_dolares_sem_ant')
    t_vad              = _safe_sum(valores, 'monto')
    t_vad_sem          = _safe_sum(ValorADepositarEmpresa.objects.all(), 'saldo_sem_ant')
    t_pf               = _safe_sum(plazos_fijos, 'monto_invertido')
    t_fci              = _safe_sum(fcis,    'saldo')
    t_fci_sem          = _safe_sum(fcis,    'saldo_sem_ant')
    t_tit_pesos        = _safe_sum(titulos, 'saldo_pesos_actual')
    t_tit_usd          = _safe_sum(titulos, 'saldo_usd_actual')
    t_tit_pesos_sem    = _safe_sum(titulos, 'saldo_pesos_sem_ant')
    t_tit_usd_sem      = _safe_sum(titulos, 'saldo_usd_sem_ant')
    t_cb_pesos         = t_caja + t_bancos_cc
    t_cb_usd           = t_bancos_usd + t_me
    t_cb_sem_pesos     = t_caja_sem + t_bancos_sem_pesos
    t_cb_sem_usd       = t_bancos_sem_usd + t_me_sem
    t_inv              = t_pf + t_fci + t_tit_pesos
    t_inv_sem          = t_fci_sem + t_tit_pesos_sem

    # ── Libro ───────────────────────────────────────────────────────────────
    wb = Workbook()

    # ════════════════════════════════════════════════════════════════════════
    # HOJA 1 — Caja y Bancos
    # ════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = 'Caja y Bancos'
    ws1.sheet_view.showGridLines = False

    # --- CAJA ---
    escribir_seccion(
        ws1, 'CAJA',
        ['Empresa', 'Saldo Sem. Anterior', 'Saldo'],
        [
            [c.empresa, fmt(c.saldo_sem_ant), fmt(c.saldo)]
            for c in cajas
        ],
        sem_ant_cols={2},
    )
    escribir_total(ws1, 3, {2: fmt(t_caja_sem), 3: fmt(t_caja)}, sem_ant_cols={2})
    ws1.append([])

    # --- BANCOS ---
    escribir_seccion(
        ws1, 'BANCOS',
        ['Banco', 'Saldo Sem. Ant. $', 'Saldo Sem. Ant. USD',
         'Saldo Cta. Cte. $', 'Chq. Pend. Acreditar', 'Chq. Pend. Débito', 'Saldo US$'],
        [
            [
                b.nombre,
                fmt(b.saldo_sem_ant_pesos), fmt(b.saldo_usd_sem_ant),
                fmt(b.saldo_cuenta_corriente),
                fmt(b.cheque_pendiente_acreditar), fmt(b.cheque_pendiente_debito),
                fmt(b.saldo_usd),
            ]
            for b in bancos
        ],
        sem_ant_cols={2, 3},
    )
    escribir_total(ws1, 7, {
        2: fmt(t_bancos_sem_pesos), 3: fmt(t_bancos_sem_usd),
        4: fmt(t_bancos_cc), 5: fmt(t_bancos_chq_acred),
        6: fmt(t_bancos_chq_deb), 7: fmt(t_bancos_usd),
    }, sem_ant_cols={2, 3})
    ws1.append([])

    # --- MONEDA EXTRANJERA ---
    escribir_seccion(
        ws1, 'MONEDA EXTRANJERA',
        ['Empresa', 'Saldo Sem. Anterior US$', 'Dólares (US$)'],
        [
            [m.empresa, fmt(m.saldo_dolares_sem_ant), fmt(m.saldo_dolares)]
            for m in monedas
        ],
        sem_ant_cols={2},
    )
    escribir_total(ws1, 3, {2: fmt(t_me_sem), 3: fmt(t_me)}, sem_ant_cols={2})

    ajustar_columnas(ws1)

    # ════════════════════════════════════════════════════════════════════════
    # HOJA 2 — Valores a Depositar
    # ════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet('Valores a Depositar')
    ws2.sheet_view.showGridLines = False

    # --- POR EMPRESA ---
    escribir_seccion(
        ws2, 'POR EMPRESA',
        ['Empresa', 'Saldo Sem. Anterior', 'Saldo'],
        [
            [row['empresa'], fmt(vad_sem_ant_map.get(row['empresa'])), fmt(row['total'])]
            for row in total_por_empresa
        ],
        sem_ant_cols={2},
    )
    escribir_total(ws2, 3, {2: fmt(t_vad_sem), 3: fmt(t_vad)}, sem_ant_cols={2})
    ws2.append([])

    # --- VENCIMIENTOS POR MES ---
    escribir_seccion(
        ws2, 'VENCIMIENTOS POR MES',
        ['Mes', 'Monto'],
        [
            [f"{row['mes_vencimiento']} {row['anio_vencimiento']}", fmt(row['total'])]
            for row in total_por_mes
        ],
    )
    escribir_total(ws2, 2, {2: fmt(t_vad)})
    ws2.append([])

    # --- DETALLE ---
    escribir_seccion(
        ws2, 'DETALLE',
        ['Empresa', 'Mes Vencimiento', 'Año', 'Monto'],
        [
            [v.empresa, v.mes_vencimiento, v.anio_vencimiento, fmt(v.monto)]
            for v in valores.order_by('empresa', 'anio_vencimiento', 'mes_vencimiento')
        ],
    )

    ajustar_columnas(ws2)

    # ════════════════════════════════════════════════════════════════════════
    # HOJA 3 — Inversiones
    # ════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet('Inversiones')
    ws3.sheet_view.showGridLines = False

    # --- PLAZOS FIJOS ---
    escribir_seccion(
        ws3, 'PLAZOS FIJOS',
        ['Banco', 'Monto Invertido', 'Fecha Constitución', 'Fecha Vencimiento', 'Interés (%)'],
        [
            [
                pf.banco,
                fmt(pf.monto_invertido),
                pf.fecha_constitucion.strftime('%d/%m/%Y') if pf.fecha_constitucion else None,
                pf.fecha_vencimiento.strftime('%d/%m/%Y') if pf.fecha_vencimiento else None,
                fmt(pf.interes),
            ]
            for pf in plazos_fijos
        ],
    )
    escribir_total(ws3, 5, {2: fmt(t_pf)})
    ws3.append([])

    # --- FCI ---
    escribir_seccion(
        ws3, 'FONDOS COMUNES DE INVERSIÓN',
        ['Nombre', 'Banco', 'Cuotap. Sem. Ant.', 'Saldo Sem. Ant.',
         'Cuotapartes Actual', 'Saldo Actual'],
        [
            [
                f.nombre, f.banco,
                fmt(f.cuotapartes_sem_ant), fmt(f.saldo_sem_ant),
                fmt(f.cuotapartes), fmt(f.saldo),
            ]
            for f in fcis
        ],
        sem_ant_cols={3, 4},
    )
    escribir_total(ws3, 6, {4: fmt(t_fci_sem), 6: fmt(t_fci)}, sem_ant_cols={3, 4})
    ws3.append([])

    # --- TÍTULOS Y ONs ---
    escribir_seccion(
        ws3, 'TÍTULOS Y ONs',
        ['Nombre', 'Tipo', 'Ticker',
         'Cuotap. Sem. Ant.', 'Saldo $ Sem. Ant.', 'Saldo USD Sem. Ant.',
         'Cuotap. Actual', 'Saldo $ Actual', 'Saldo USD Actual'],
        [
            [
                t.nombre, t.tipo, t.ticker,
                fmt(t.cuotapartes_sem_ant), fmt(t.saldo_pesos_sem_ant), fmt(t.saldo_usd_sem_ant),
                fmt(t.cuotapartes_actual), fmt(t.saldo_pesos_actual), fmt(t.saldo_usd_actual),
            ]
            for t in titulos
        ],
        sem_ant_cols={4, 5, 6},
    )
    escribir_total(ws3, 9, {
        5: fmt(t_tit_pesos_sem), 6: fmt(t_tit_usd_sem),
        8: fmt(t_tit_pesos), 9: fmt(t_tit_usd),
    }, sem_ant_cols={4, 5, 6})

    ajustar_columnas(ws3)

    # ════════════════════════════════════════════════════════════════════════
    # HOJA 4 — Resumen General
    # ════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet('Resumen General')
    ws4.sheet_view.showGridLines = False

    headers_resumen = ['Concepto', 'Sem. Ant. $', 'Pesos ($)', 'Sem. Ant. US$', 'Dólares (US$)']
    ws4.append(headers_resumen)
    fila_cab = ws4.max_row
    for col_idx, _ in enumerate(headers_resumen, 1):
        c = ws4.cell(fila_cab, col_idx)
        c.font = font_header
        c.fill = fill_header
        c.alignment = aln_center
        c.border = border
        if col_idx in {2, 4}:
            c.fill = PatternFill('solid', fgColor='6B3A8A')  # púrpura medio para sem_ant

    def fila_resumen(ws, concepto, sem_pesos, pesos, sem_usd, usd):
        ws.append([concepto, sem_pesos, pesos, sem_usd, usd])
        r = ws.max_row
        for col_idx in range(1, 6):
            c = ws.cell(r, col_idx)
            c.font = font_normal
            c.border = border
            v = c.value
            if col_idx in {2, 4}:
                c.fill = fill_sem_ant
            if isinstance(v, float):
                c.alignment = aln_right
                c.number_format = '#,##0.00'
            else:
                c.alignment = aln_left

    def fila_total_resumen(ws, concepto, sem_pesos, pesos, sem_usd, usd):
        ws.append([concepto, sem_pesos, pesos, sem_usd, usd])
        r = ws.max_row
        for col_idx in range(1, 6):
            c = ws.cell(r, col_idx)
            c.font = font_total
            c.fill = fill_total
            c.border = border
            if col_idx in {2, 4}:
                c.fill = PatternFill('solid', fgColor='D5C8E0')
            v = c.value
            if isinstance(v, float):
                c.alignment = aln_right
                c.number_format = '#,##0.00'
            else:
                c.alignment = aln_left

    fila_resumen(ws4, 'Caja',               fmt(t_caja_sem),     fmt(t_caja),     None,              None)
    fila_resumen(ws4, 'Bancos',             fmt(t_bancos_sem_pesos), fmt(t_bancos_cc), fmt(t_bancos_sem_usd), fmt(t_bancos_usd))
    fila_resumen(ws4, 'Moneda Extranjera',  None,                None,            fmt(t_me_sem),     fmt(t_me))
    fila_resumen(ws4, 'Valores a Depositar', fmt(t_vad_sem),     fmt(t_vad),      None,              None)
    fila_total_resumen(ws4, 'CAJA Y BANCOS TOTAL', fmt(t_cb_sem_pesos), fmt(t_cb_pesos), fmt(t_cb_sem_usd), fmt(t_cb_usd))

    # Separador visual
    ws4.append([])

    fila_resumen(ws4, 'Plazos Fijos',  None,           fmt(t_pf),      None,               None)
    fila_resumen(ws4, 'FCI',           fmt(t_fci_sem), fmt(t_fci),     None,               None)
    fila_resumen(ws4, 'Títulos / ONs', fmt(t_tit_pesos_sem), fmt(t_tit_pesos), fmt(t_tit_usd_sem), fmt(t_tit_usd))
    fila_total_resumen(ws4, 'INVERSIONES TOTAL', fmt(t_inv_sem), fmt(t_inv), fmt(t_tit_usd_sem), fmt(t_tit_usd))

    ajustar_columnas(ws4)

    # ── Respuesta ────────────────────────────────────────────────────────────
    from datetime import date as _date
    nombre_archivo = f"tesoreria_{_date.today().strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)
    return response


@login_required
def caja_crear(request):
    _check_rol(request.user)
    if request.method == 'POST':
        form = CajaForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('tesoreria:caja_bancos')
    else:
        form = CajaForm()
    return render(request, 'tesoreria/form_generic.html', {
        'form': form, 'titulo': 'Nueva Caja', 'vista_activa': 'caja_bancos',
    })

@login_required
def caja_editar(request, pk):
    _check_rol(request.user)
    obj = get_object_or_404(Caja, pk=pk)
    if request.method == 'POST':
        form = CajaForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            return redirect('tesoreria:caja_bancos')
    else:
        form = CajaForm(instance=obj)
    return render(request, 'tesoreria/form_generic.html', {
        'form': form, 'titulo': 'Editar Caja', 'vista_activa': 'caja_bancos',
    })

@login_required
def banco_crear(request):
    _check_rol(request.user)
    if request.method == 'POST':
        form = BancoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('tesoreria:caja_bancos')
    else:
        form = BancoForm()
    return render(request, 'tesoreria/form_generic.html', {
        'form': form, 'titulo': 'Nuevo Banco', 'vista_activa': 'caja_bancos',
    })

@login_required
def banco_editar(request, pk):
    _check_rol(request.user)
    obj = get_object_or_404(Banco, pk=pk)
    if request.method == 'POST':
        form = BancoForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            return redirect('tesoreria:caja_bancos')
    else:
        form = BancoForm(instance=obj)
    return render(request, 'tesoreria/form_generic.html', {
        'form': form, 'titulo': f'Editar Banco: {obj.nombre}', 'vista_activa': 'caja_bancos',
    })

@login_required
def banco_eliminar(request, pk):
    _check_rol(request.user)
    obj = get_object_or_404(Banco, pk=pk)
    if request.method == 'POST':
        obj.delete()
        return redirect('tesoreria:caja_bancos')
    return render(request, 'tesoreria/confirm_delete.html', {
        'objeto': obj, 'titulo': 'Eliminar Banco', 'vista_activa': 'caja_bancos',
    })

@login_required
def moneda_extranjera_crear(request):
    _check_rol(request.user)
    if request.method == 'POST':
        form = MonedaExtranjeraForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('tesoreria:caja_bancos')
    else:
        form = MonedaExtranjeraForm()
    return render(request, 'tesoreria/form_generic.html', {
        'form': form, 'titulo': 'Nueva Moneda Extranjera', 'vista_activa': 'caja_bancos',
    })

@login_required
def moneda_extranjera_editar(request, pk):
    _check_rol(request.user)
    obj = get_object_or_404(MonedaExtranjera, pk=pk)
    if request.method == 'POST':
        form = MonedaExtranjeraForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            return redirect('tesoreria:caja_bancos')
    else:
        form = MonedaExtranjeraForm(instance=obj)
    return render(request, 'tesoreria/form_generic.html', {
        'form': form, 'titulo': f'Editar Moneda Extranjera: {obj.empresa}',
        'vista_activa': 'caja_bancos',
    })

@login_required
def moneda_extranjera_eliminar(request, pk):
    _check_rol(request.user)
    obj = get_object_or_404(MonedaExtranjera, pk=pk)
    if request.method == 'POST':
        obj.delete()
        return redirect('tesoreria:caja_bancos')
    return render(request, 'tesoreria/confirm_delete.html', {
        'objeto': obj, 'titulo': 'Eliminar Moneda Extranjera', 'vista_activa': 'caja_bancos',
    })

@login_required
def vad_crear(request):
    _check_rol(request.user)
    if request.method == 'POST':
        form = ValorADepositarForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('tesoreria:valores_a_depositar')
    else:
        form = ValorADepositarForm()
    return render(request, 'tesoreria/form_generic.html', {
        'form': form, 'titulo': 'Nuevo Valor a Depositar',
        'vista_activa': 'valores_a_depositar',
    })

@login_required
def vad_editar(request, pk):
    _check_rol(request.user)
    obj = get_object_or_404(ValorADepositar, pk=pk)
    if request.method == 'POST':
        form = ValorADepositarForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            return redirect('tesoreria:valores_a_depositar')
    else:
        form = ValorADepositarForm(instance=obj)
    return render(request, 'tesoreria/form_generic.html', {
        'form': form, 'titulo': 'Editar Valor a Depositar',
        'vista_activa': 'valores_a_depositar',
    })

@login_required
def vad_eliminar(request, pk):
    _check_rol(request.user)
    obj = get_object_or_404(ValorADepositar, pk=pk)
    if request.method == 'POST':
        obj.delete()
        return redirect('tesoreria:valores_a_depositar')
    return render(request, 'tesoreria/confirm_delete.html', {
        'objeto': obj, 'titulo': 'Eliminar Valor a Depositar',
        'vista_activa': 'valores_a_depositar',
    })

@login_required
def pf_crear(request):
    _check_rol(request.user)
    if request.method == 'POST':
        form = PlazoFijoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('tesoreria:inversiones')
    else:
        form = PlazoFijoForm()
    return render(request, 'tesoreria/form_generic.html', {
        'form': form, 'titulo': 'Nuevo Plazo Fijo', 'vista_activa': 'inversiones',
    })

@login_required
def pf_editar(request, pk):
    _check_rol(request.user)
    obj = get_object_or_404(PlazoFijo, pk=pk)
    if request.method == 'POST':
        form = PlazoFijoForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            return redirect('tesoreria:inversiones')
    else:
        form = PlazoFijoForm(instance=obj)
    return render(request, 'tesoreria/form_generic.html', {
        'form': form, 'titulo': f'Editar Plazo Fijo: {obj.banco}',
        'vista_activa': 'inversiones',
    })

@login_required
def pf_eliminar(request, pk):
    _check_rol(request.user)
    obj = get_object_or_404(PlazoFijo, pk=pk)
    if request.method == 'POST':
        obj.delete()
        return redirect('tesoreria:inversiones')
    return render(request, 'tesoreria/confirm_delete.html', {
        'objeto': obj, 'titulo': 'Eliminar Plazo Fijo',
        'vista_activa': 'inversiones',
    })

@login_required
def fci_crear(request):
    _check_rol(request.user)
    if request.method == 'POST':
        form = FCIForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('tesoreria:inversiones')
    else:
        form = FCIForm()
    return render(request, 'tesoreria/form_generic.html', {
        'form': form, 'titulo': 'Nuevo FCI', 'vista_activa': 'inversiones',
    })

@login_required
def fci_editar(request, pk):
    _check_rol(request.user)
    obj = get_object_or_404(FCI, pk=pk)
    if request.method == 'POST':
        form = FCIForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            return redirect('tesoreria:inversiones')
    else:
        form = FCIForm(instance=obj)
    return render(request, 'tesoreria/form_generic.html', {
        'form': form, 'titulo': f'Editar FCI: {obj.nombre}',
        'vista_activa': 'inversiones',
    })

@login_required
def fci_eliminar(request, pk):
    _check_rol(request.user)
    obj = get_object_or_404(FCI, pk=pk)
    if request.method == 'POST':
        obj.delete()
        return redirect('tesoreria:inversiones')
    return render(request, 'tesoreria/confirm_delete.html', {
        'objeto': obj, 'titulo': 'Eliminar FCI', 'vista_activa': 'inversiones',
    })

@login_required
def titulo_crear(request):
    _check_rol(request.user)
    if request.method == 'POST':
        form = TituloONForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('tesoreria:inversiones')
    else:
        form = TituloONForm()
    return render(request, 'tesoreria/form_generic.html', {
        'form': form, 'titulo': 'Nuevo Título / ON', 'vista_activa': 'inversiones',
    })

@login_required
def titulo_editar(request, pk):
    _check_rol(request.user)
    obj = get_object_or_404(TituloON, pk=pk)
    if request.method == 'POST':
        form = TituloONForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            return redirect('tesoreria:inversiones')
    else:
        form = TituloONForm(instance=obj)
    return render(request, 'tesoreria/form_generic.html', {
        'form': form, 'titulo': f'Editar Título: {obj.nombre}',
        'vista_activa': 'inversiones',
    })

@login_required
def titulo_eliminar(request, pk):
    _check_rol(request.user)
    obj = get_object_or_404(TituloON, pk=pk)
    if request.method == 'POST':
        obj.delete()
        return redirect('tesoreria:inversiones')
    return render(request, 'tesoreria/confirm_delete.html', {
        'objeto': obj, 'titulo': 'Eliminar Título / ON',
        'vista_activa': 'inversiones',
    })