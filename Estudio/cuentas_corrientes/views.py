import json
import logging
from datetime import date
from decimal import Decimal, InvalidOperation
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q, Sum, F, Value, DecimalField
from django.db.models.functions import Coalesce
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.views.decorators.http import require_POST
from openpyxl import load_workbook, Workbook
from .models import ClienteCC, MesCC, ConfiguracionMeses

logger = logging.getLogger(__name__)

ROLES_PERMITIDOS = ('Administrador', 'Colaborador')
MESES_ES = {
    1: 'Ene', 2: 'Feb', 3: 'Mar', 4: 'Abr',
    5: 'May', 6: 'Jun', 7: 'Jul', 8: 'Ago',
    9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dic',
}
MESES_ES_REVERSE = {v: k for k, v in MESES_ES.items()}

def _check_rol(user):
    if user.rol not in ROLES_PERMITIDOS:
        raise PermissionDenied

def _get_periodos_activos():
    return list(ConfiguracionMeses.objects.order_by('orden').values_list('periodo', flat=True))

def _limpiar_meses_inactivos():
    periodos_vigentes = list(ConfiguracionMeses.objects.values_list('periodo', flat=True))
    sumas = (
        MesCC.objects
        .exclude(periodo__in=periodos_vigentes)
        .values('cliente_id')
        .annotate(total=Sum('monto'))
    )
    for item in sumas:
        ClienteCC.objects.filter(pk=item['cliente_id']).update(
            vencido=F('vencido') + item['total']
        )
    MesCC.objects.exclude(periodo__in=periodos_vigentes).delete()

def _formato_periodo(periodo):
    return f'{MESES_ES.get(periodo.month, "?")}-{str(periodo.year)[2:]}'

def _get_dolar_mep_venta():
    try:
        from cotizaciones.services import get_dolares
        dolares = get_dolares()
        for d in dolares:
            if d.get('casa') == 'bolsa':
                venta = d.get('venta')
                if venta is not None:
                    return Decimal(str(venta))
    except Exception:
        logger.warning('No se pudo obtener el dólar MEP.')
    return None

def _build_tabla(clientes, periodos_activos):
    filas = []
    for cliente in clientes:
        meses_dict = {}
        for mes in cliente.meses.filter(periodo__in=periodos_activos):
            meses_dict[mes.periodo] = mes.monto

        meses_valores = []
        for p in periodos_activos:
            meses_valores.append({
                'periodo': p.isoformat(),
                'monto': meses_dict.get(p, Decimal('0.00')),
            })

        suma_meses = sum(m['monto'] for m in meses_valores)
        saldo = cliente.vencido + cliente.balance_especial + suma_meses

        # Morosidad: (vencido + mes_orden1 + mes_orden2) / (saldo - balance_especial) * 100
        if saldo == 0:
            morosidad = Decimal('0.00')
        else:
            denominador = saldo - cliente.balance_especial  # = vencido + suma_meses
            monto_orden1 = meses_valores[0]['monto'] if len(meses_valores) > 0 else Decimal('0.00')
            monto_orden2 = meses_valores[1]['monto'] if len(meses_valores) > 1 else Decimal('0.00')
            numerador = cliente.vencido + monto_orden1 + monto_orden2
            morosidad = (numerador / denominador * 100) if denominador != 0 else None

        filas.append({
            'cliente': cliente,
            'saldo': saldo,
            'vencido': cliente.vencido,
            'balance_especial': cliente.balance_especial,
            'meses': meses_valores,
            'morosidad': morosidad,
        })
    return filas

@login_required
def lista_cuentas(request):
    _check_rol(request.user)

    periodos_activos = _get_periodos_activos()
    encabezados_meses = [_formato_periodo(p) for p in periodos_activos]

    # Filtro por nombre
    busqueda = request.GET.get('busqueda', '').strip()
    clientes = ClienteCC.objects.filter(activo=True)
    if busqueda:
        clientes = clientes.filter(nombre__icontains=busqueda)

    # Anotar suma de meses activos para poder ordenar por saldo en DB
    anotaciones_meses = {}
    if periodos_activos:
        clientes = clientes.annotate(
            suma_meses=Coalesce(
                Sum('meses__monto', filter=Q(meses__periodo__in=periodos_activos)),
                Value(Decimal('0.00')),
                output_field=DecimalField(max_digits=15, decimal_places=2),
            ),
            saldo_calculado=F('vencido') + F('balance_especial') + F('suma_meses'),
        )
        # Anotar cada periodo individual para poder ordenar por mes
        for idx, periodo in enumerate(periodos_activos):
            alias = f'mes_{idx}'
            clientes = clientes.annotate(**{
                alias: Coalesce(
                    Sum('meses__monto', filter=Q(meses__periodo=periodo)),
                    Value(Decimal('0.00')),
                    output_field=DecimalField(max_digits=15, decimal_places=2),
                ),
            })
            anotaciones_meses[idx] = alias
    else:
        clientes = clientes.annotate(
            suma_meses=Value(Decimal('0.00'), output_field=DecimalField(max_digits=15, decimal_places=2)),
            saldo_calculado=F('vencido') + F('balance_especial'),
        )

    # Ordenamiento — mutuamente excluyente
    orden = request.GET.get('orden', '')
    orden_aplicado = False
    ordenes_validos = {
        'saldo_asc': 'saldo_calculado',
        'saldo_desc': '-saldo_calculado',
        'vencido_asc': 'vencido',
        'vencido_desc': '-vencido',
        'balance_asc': 'balance_especial',
        'balance_desc': '-balance_especial',
    }
    # Ordenes dinámicos por mes
    for idx in range(len(periodos_activos)):
        alias = anotaciones_meses.get(idx, f'mes_{idx}')
        ordenes_validos[f'mes{idx}_asc'] = alias
        ordenes_validos[f'mes{idx}_desc'] = f'-{alias}'

    if orden in ordenes_validos:
        clientes = clientes.order_by(ordenes_validos[orden])
        orden_aplicado = True

    if not orden_aplicado:
        clientes = clientes.order_by('nombre')

    # Paginación
    paginator = Paginator(clientes, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    filas = _build_tabla(page_obj.object_list, periodos_activos)

    # Calcular totales por columna (de la página actual)
    totales = {
        'saldo': sum(f['saldo'] for f in filas),
        'vencido': sum(f['vencido'] for f in filas),
        'balance_especial': sum(f['balance_especial'] for f in filas),
        'meses': [
            sum(f['meses'][i]['monto'] for f in filas)
            for i in range(len(periodos_activos))
        ] if periodos_activos and filas else [],
    }

    # ── Tablero de control: totales GLOBALES (todos los clientes, no solo la página) ──
    todos_clientes = ClienteCC.objects.filter(activo=True)
    if periodos_activos:
        agg = todos_clientes.aggregate(
            total_vencido=Coalesce(
                Sum('vencido'), Value(Decimal('0.00')),
                output_field=DecimalField(max_digits=15, decimal_places=2),
            ),
            total_balance=Coalesce(
                Sum('balance_especial'), Value(Decimal('0.00')),
                output_field=DecimalField(max_digits=15, decimal_places=2),
            ),
            total_meses=Coalesce(
                Sum('meses__monto', filter=Q(meses__periodo__in=periodos_activos)),
                Value(Decimal('0.00')),
                output_field=DecimalField(max_digits=15, decimal_places=2),
            ),
        )
    else:
        agg = todos_clientes.aggregate(
            total_vencido=Coalesce(
                Sum('vencido'), Value(Decimal('0.00')),
                output_field=DecimalField(max_digits=15, decimal_places=2),
            ),
            total_balance=Coalesce(
                Sum('balance_especial'), Value(Decimal('0.00')),
                output_field=DecimalField(max_digits=15, decimal_places=2),
            ),
        )
        agg['total_meses'] = Decimal('0.00')

    dashboard = {
        'total_saldo': agg['total_vencido'] + agg['total_balance'] + agg['total_meses'],
        'total_vencido': agg['total_vencido'],
        'total_balance': agg['total_balance'],
    }

    # Obtener dólar MEP/bolsa (venta) desde cotizaciones
    dolar_venta = _get_dolar_mep_venta()
    dashboard['dolar_venta'] = dolar_venta
    if dolar_venta and dolar_venta > 0:
        dashboard['total_saldo_usd'] = dashboard['total_saldo'] / dolar_venta
        dashboard['total_vencido_usd'] = dashboard['total_vencido'] / dolar_venta
        dashboard['total_balance_usd'] = dashboard['total_balance'] / dolar_venta
    else:
        dashboard['total_saldo_usd'] = None
        dashboard['total_vencido_usd'] = None
        dashboard['total_balance_usd'] = None

    # Sumatoria estática de facturación del último mes creado
    ultimo_periodo_config = ConfiguracionMeses.objects.order_by('-orden').first()
    dashboard['sumatoria_facturacion'] = (
        ultimo_periodo_config.sumatoria_facturacion
        if ultimo_periodo_config else Decimal('0.00')
    )
    dashboard['periodo_facturacion'] = (
        _formato_periodo(ultimo_periodo_config.periodo)
        if ultimo_periodo_config else '—'
    )

    contexto = {
        'filas': filas,
        'encabezados_meses': encabezados_meses,
        'periodos_activos': periodos_activos,
        'busqueda': busqueda,
        'orden': orden,
        'page_obj': page_obj,
        'totales': totales,
        'dashboard': dashboard,
    }
    return render(request, 'cuentas_corrientes/lista.html', contexto)


@login_required
@require_POST
def editar_fila(request):
    _check_rol(request.user)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'JSON inválido.'}, status=400)

    cliente_id = data.get('cliente_id')
    if not cliente_id:
        return JsonResponse({'error': 'Falta cliente_id.'}, status=400)

    cliente = get_object_or_404(ClienteCC, pk=cliente_id, activo=True)
    periodos_activos = _get_periodos_activos()

    try:
        vencido = Decimal(str(data.get('vencido', cliente.vencido)))
        balance_especial = Decimal(str(data.get('balance_especial', cliente.balance_especial)))
    except (InvalidOperation, TypeError):
        return JsonResponse({'error': 'Valores numéricos inválidos.'}, status=400)

    meses_data = data.get('meses', {})

    with transaction.atomic():
        cliente.vencido = vencido
        cliente.balance_especial = balance_especial
        cliente.save(update_fields=['vencido', 'balance_especial'])

        for periodo_str, monto_str in meses_data.items():
            try:
                periodo = date.fromisoformat(periodo_str)
                monto = Decimal(str(monto_str))
            except (ValueError, InvalidOperation):
                continue

            if periodo in periodos_activos:
                MesCC.objects.update_or_create(
                    cliente=cliente,
                    periodo=periodo,
                    defaults={'monto': monto},
                )

    # Recalcular saldo
    suma_meses = cliente.meses.filter(
        periodo__in=periodos_activos
    ).aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
    nuevo_saldo = cliente.vencido + cliente.balance_especial + suma_meses

    # Devolver los meses actualizados
    meses_resp = {}
    for mes in cliente.meses.filter(periodo__in=periodos_activos):
        meses_resp[mes.periodo.isoformat()] = str(mes.monto)

    # Recalcular totales globales para el dashboard
    todos = ClienteCC.objects.filter(activo=True)
    if periodos_activos:
        agg = todos.aggregate(
            total_vencido=Coalesce(
                Sum('vencido'), Value(Decimal('0.00')),
                output_field=DecimalField(max_digits=15, decimal_places=2),
            ),
            total_balance=Coalesce(
                Sum('balance_especial'), Value(Decimal('0.00')),
                output_field=DecimalField(max_digits=15, decimal_places=2),
            ),
            total_meses=Coalesce(
                Sum('meses__monto', filter=Q(meses__periodo__in=periodos_activos)),
                Value(Decimal('0.00')),
                output_field=DecimalField(max_digits=15, decimal_places=2),
            ),
        )
    else:
        agg = todos.aggregate(
            total_vencido=Coalesce(
                Sum('vencido'), Value(Decimal('0.00')),
                output_field=DecimalField(max_digits=15, decimal_places=2),
            ),
            total_balance=Coalesce(
                Sum('balance_especial'), Value(Decimal('0.00')),
                output_field=DecimalField(max_digits=15, decimal_places=2),
            ),
        )
        agg['total_meses'] = Decimal('0.00')

    dashboard_totals = {
        'total_saldo': str(agg['total_vencido'] + agg['total_balance'] + agg['total_meses']),
        'total_vencido': str(agg['total_vencido']),
        'total_balance': str(agg['total_balance']),
    }

    return JsonResponse({
        'ok': True,
        'saldo': str(nuevo_saldo),
        'vencido': str(cliente.vencido),
        'balance_especial': str(cliente.balance_especial),
        'meses': meses_resp,
        'dashboard': dashboard_totals,
    })

@login_required
def nuevo_mes(request):
    _check_rol(request.user)
    resultado = None

    if request.method == 'POST' and request.FILES.get('archivo'):
        archivo = request.FILES['archivo']

        if not archivo.name.endswith('.xlsx'):
            resultado = {'error': 'El archivo debe ser un .xlsx'}
            return render(request, 'cuentas_corrientes/nuevo_mes.html', {'resultado': resultado})

        try:
            wb = load_workbook(archivo, read_only=True)
            ws = wb.active
        except Exception:
            resultado = {'error': 'No se pudo leer el archivo Excel.'}
            return render(request, 'cuentas_corrientes/nuevo_mes.html', {'resultado': resultado})

        # Leer datos del Excel
        datos_excel = []
        errores_formato = []
        for idx, row in enumerate(ws.iter_rows(min_row=2, max_col=2, values_only=True), start=2):
            nombre_raw, monto_raw = row
            if nombre_raw is None:
                continue
            nombre = str(nombre_raw).strip()
            if not nombre:
                continue
            try:
                monto = Decimal(str(monto_raw)) if monto_raw is not None else Decimal('0.00')
            except (InvalidOperation, ValueError):
                errores_formato.append(f'Fila {idx}: monto inválido para "{nombre}"')
                continue
            datos_excel.append((nombre, monto))

        if not datos_excel and not errores_formato:
            resultado = {'error': 'El archivo está vacío o no tiene datos válidos.'}
            return render(request, 'cuentas_corrientes/nuevo_mes.html', {'resultado': resultado})

        periodos_activos = list(
            ConfiguracionMeses.objects.order_by('orden').values('orden', 'periodo')
        )

        if len(periodos_activos) < 5:
            resultado = {
                'error': (
                    'La configuración de meses activos está incompleta. '
                    'Deben existir exactamente 5 periodos en ConfiguracionMeses. '
                    f'Actualmente hay {len(periodos_activos)}.'
                )
            }
            return render(request, 'cuentas_corrientes/nuevo_mes.html', {'resultado': resultado})

        hoy = date.today()
        periodo_corriente = date(hoy.year, hoy.month, 1)
        periodo_mas_reciente = periodos_activos[-1]['periodo']

        # Determinar si la columna del mes corriente ya existe
        mes_corriente_existe = any(
            p['periodo'] == periodo_corriente for p in periodos_activos
        )

        procesados = 0
        clientes_creados = 0

        if mes_corriente_existe:
            # Escenario 2: la columna del mes corriente ya existe 
            # Solo sumar facturación a la columna existente.
            with transaction.atomic():
                for nombre, monto in datos_excel:
                    try:
                        cliente = ClienteCC.objects.get(nombre__iexact=nombre)
                    except ClienteCC.DoesNotExist:
                        cliente = ClienteCC.objects.create(nombre=nombre)
                        clientes_creados += 1

                    mes_existente = MesCC.objects.filter(
                        cliente=cliente,
                        periodo=periodo_corriente,
                    ).first()
                    if mes_existente:
                        mes_existente.monto += monto
                        mes_existente.save(update_fields=['monto'])
                    else:
                        MesCC.objects.create(
                            cliente=cliente,
                            periodo=periodo_corriente,
                            monto=monto,
                        )
                    procesados += 1
                # Limpiar huérfanos por si existieran periodos inactivos no eliminados
                _limpiar_meses_inactivos()

                # Actualizar sumatoria estática de facturación
                sumatoria = MesCC.objects.filter(
                    periodo=periodo_corriente
                ).aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
                ConfiguracionMeses.objects.filter(
                    periodo=periodo_corriente
                ).update(sumatoria_facturacion=sumatoria)

            resultado = {
                'ok': True,
                'modo': 'actualizacion',
                'procesados': procesados,
                'clientes_creados': clientes_creados,
                'errores_formato': errores_formato,
                'periodo_actualizado': _formato_periodo(periodo_corriente),
            }
        else:
            # ── Escenario 1: primera vez en el mes ──
            # Validar que el próximo periodo coincida con el mes corriente.
            if periodo_mas_reciente.month == 12:
                nuevo_periodo = date(periodo_mas_reciente.year + 1, 1, 1)
            else:
                nuevo_periodo = date(periodo_mas_reciente.year, periodo_mas_reciente.month + 1, 1)

            if nuevo_periodo != periodo_corriente:
                resultado = {
                    'error': (
                        f'El mes a crear ({_formato_periodo(nuevo_periodo)}) no coincide con el mes corriente '
                        f'({_formato_periodo(periodo_corriente)}). '
                        'El proceso de Nuevo Mes solo puede ejecutarse durante el mes correspondiente.'
                    )
                }
                return render(request, 'cuentas_corrientes/nuevo_mes.html', {'resultado': resultado})

            periodo_mas_antiguo = periodos_activos[0]['periodo']

            with transaction.atomic():
                ConfiguracionMeses.objects.filter(periodo=periodo_mas_antiguo).delete()
                for config in ConfiguracionMeses.objects.order_by('orden'):
                    if config.orden > 1:
                        config.orden -= 1
                        config.save(update_fields=['orden'])
                ConfiguracionMeses.objects.create(orden=5, periodo=nuevo_periodo)
                _limpiar_meses_inactivos()
                for nombre, monto in datos_excel:
                    try:
                        cliente = ClienteCC.objects.get(nombre__iexact=nombre)
                    except ClienteCC.DoesNotExist:
                        cliente = ClienteCC.objects.create(nombre=nombre)
                        clientes_creados += 1
                    MesCC.objects.update_or_create(
                        cliente=cliente,
                        periodo=nuevo_periodo,
                        defaults={'monto': monto},
                    )
                    procesados += 1

                sumatoria = MesCC.objects.filter(
                    periodo=nuevo_periodo
                ).aggregate(total=Sum('monto'))['total'] or Decimal('0.00')
                ConfiguracionMeses.objects.filter(
                    periodo=nuevo_periodo
                ).update(sumatoria_facturacion=sumatoria)

            resultado = {
                'ok': True,
                'modo': 'creacion',
                'procesados': procesados,
                'clientes_creados': clientes_creados,
                'errores_formato': errores_formato,
                'periodo_cerrado': _formato_periodo(periodo_mas_antiguo),
                'periodo_nuevo': _formato_periodo(nuevo_periodo),
            }

    return render(request, 'cuentas_corrientes/nuevo_mes.html', {'resultado': resultado})

@login_required
def exportar_excel(request):
    _check_rol(request.user)

    periodos_activos = _get_periodos_activos()
    encabezados_meses = [_formato_periodo(p) for p in periodos_activos]
    busqueda = request.GET.get('busqueda', '').strip()
    clientes = ClienteCC.objects.filter(activo=True)
    if busqueda:
        clientes = clientes.filter(nombre__icontains=busqueda)

    if periodos_activos:
        clientes = clientes.annotate(
            suma_meses=Coalesce(
                Sum('meses__monto', filter=Q(meses__periodo__in=periodos_activos)),
                Value(Decimal('0.00')),
                output_field=DecimalField(max_digits=15, decimal_places=2),
            ),
            saldo_calculado=F('vencido') + F('balance_especial') + F('suma_meses'),
        )
    else:
        clientes = clientes.annotate(
            suma_meses=Value(Decimal('0.00'), output_field=DecimalField(max_digits=15, decimal_places=2)),
            saldo_calculado=F('vencido') + F('balance_especial'),
        )

    orden = request.GET.get('orden', '')
    if orden == 'saldo_asc':
        clientes = clientes.order_by('saldo_calculado')
    elif orden == 'saldo_desc':
        clientes = clientes.order_by('-saldo_calculado')
    else:
        clientes = clientes.order_by('nombre')

    filas = _build_tabla(clientes, periodos_activos)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Cuentas Corrientes'

    # Encabezados
    headers = ['Cliente', 'Saldo', 'Vencido', 'Balance/Especial'] + encabezados_meses
    ws.append(headers)

    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='8C4F9F', end_color='8C4F9F', fill_type='solid')
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for fila in filas:
        row = [
            fila['cliente'].nombre,
            float(fila['saldo']),
            float(fila['vencido']),
            float(fila['balance_especial']),
        ] + [float(m['monto']) for m in fila['meses']]
        ws.append(row)

    from openpyxl.utils import get_column_letter
    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(header) + 4, 14)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="cuentas_corrientes.xlsx"'
    wb.save(response)
    return response


def _check_admin_staff(user):
    if not (user.rol == 'Administrador' and user.is_staff):
        raise PermissionDenied


def _parse_periodo_header(header_str):
    header_str = str(header_str).strip()
    parts = header_str.split('-')
    if len(parts) != 2:
        return None
    mes_str, year_str = parts
    mes = MESES_ES_REVERSE.get(mes_str)
    if mes is None:
        return None
    try:
        year = int(year_str)
        if year < 100:
            year += 2000
        return date(year, mes, 1)
    except (ValueError, TypeError):
        return None


@login_required
def importar_excel(request):
    _check_admin_staff(request.user)
    resultado = None

    if request.method == 'POST' and request.FILES.get('archivo'):
        archivo = request.FILES['archivo']

        if not archivo.name.endswith('.xlsx'):
            resultado = {'error': 'El archivo debe ser un .xlsx'}
            return render(request, 'cuentas_corrientes/importar.html', {'resultado': resultado})

        try:
            wb = load_workbook(archivo, read_only=True)
            ws = wb.active
        except Exception:
            resultado = {'error': 'No se pudo leer el archivo Excel.'}
            return render(request, 'cuentas_corrientes/importar.html', {'resultado': resultado})

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            resultado = {'error': 'El archivo está vacío.'}
            return render(request, 'cuentas_corrientes/importar.html', {'resultado': resultado})

        headers = [str(h).strip() if h else '' for h in rows[0]]

        if len(headers) < 4:
            resultado = {
                'error': (
                    'El archivo debe tener al menos 4 columnas: '
                    'Cliente, Saldo, Vencido, Balance/Especial.'
                )
            }
            return render(request, 'cuentas_corrientes/importar.html', {'resultado': resultado})

        periodos_excel = []
        for i, h in enumerate(headers[4:], start=4):
            periodo = _parse_periodo_header(h)
            if periodo:
                periodos_excel.append((i, periodo))

        datos_excel = []
        errores_formato = []
        for idx, row in enumerate(rows[1:], start=2):
            if not row or row[0] is None:
                continue
            nombre = str(row[0]).strip()
            if not nombre:
                continue

            try:
                vencido = Decimal(str(row[2])) if row[2] is not None else Decimal('0.00')
            except (InvalidOperation, ValueError, IndexError):
                errores_formato.append(f'Fila {idx}: valor de Vencido inválido para "{nombre}".')
                continue

            try:
                balance_especial = Decimal(str(row[3])) if row[3] is not None else Decimal('0.00')
            except (InvalidOperation, ValueError, IndexError):
                errores_formato.append(f'Fila {idx}: valor de Balance/Especial inválido para "{nombre}".')
                continue

            meses = []
            for col_idx, periodo in periodos_excel:
                try:
                    val = row[col_idx] if col_idx < len(row) else None
                    monto = Decimal(str(val)) if val is not None else Decimal('0.00')
                except (InvalidOperation, ValueError):
                    errores_formato.append(
                        f'Fila {idx}: monto inválido en columna "{headers[col_idx]}" para "{nombre}".'
                    )
                    monto = Decimal('0.00')
                meses.append((periodo, monto))

            datos_excel.append({
                'nombre': nombre,
                'vencido': vencido,
                'balance_especial': balance_especial,
                'meses': meses,
            })

        if not datos_excel and not errores_formato:
            resultado = {'error': 'El archivo no contiene datos válidos.'}
            return render(request, 'cuentas_corrientes/importar.html', {'resultado': resultado})

        clientes_creados = 0
        clientes_actualizados = 0
        meses_creados = 0
        periodos_configurados = 0

        with transaction.atomic():
            if periodos_excel:
                periodos_existentes = set(
                    ConfiguracionMeses.objects.values_list('periodo', flat=True)
                )
                periodos_a_crear = [
                    p for _, p in periodos_excel if p not in periodos_existentes
                ]
                if periodos_a_crear:
                    # Determinar el orden máximo actual
                    max_orden = ConfiguracionMeses.objects.count()
                    for p in sorted(periodos_a_crear):
                        max_orden += 1
                        ConfiguracionMeses.objects.create(orden=max_orden, periodo=p)
                        periodos_configurados += 1

            for dato in datos_excel:
                cliente, creado = ClienteCC.objects.get_or_create(
                    nombre__iexact=dato['nombre'],
                    defaults={
                        'nombre': dato['nombre'],
                        'vencido': dato['vencido'],
                        'balance_especial': dato['balance_especial'],
                    },
                )
                if creado:
                    clientes_creados += 1
                else:
                    cliente.vencido = dato['vencido']
                    cliente.balance_especial = dato['balance_especial']
                    cliente.save(update_fields=['vencido', 'balance_especial'])
                    clientes_actualizados += 1

                for periodo, monto in dato['meses']:
                    _, mes_creado = MesCC.objects.update_or_create(
                        cliente=cliente,
                        periodo=periodo,
                        defaults={'monto': monto},
                    )
                    if mes_creado:
                        meses_creados += 1

        resultado = {
            'ok': True,
            'clientes_creados': clientes_creados,
            'clientes_actualizados': clientes_actualizados,
            'meses_creados': meses_creados,
            'periodos_configurados': periodos_configurados,
            'errores_formato': errores_formato,
            'total_procesados': clientes_creados + clientes_actualizados,
        }

        messages.success(
            request,
            f'Importación completada: {clientes_creados} clientes creados, '
            f'{clientes_actualizados} actualizados.',
        )

    return render(request, 'cuentas_corrientes/importar.html', {'resultado': resultado})
